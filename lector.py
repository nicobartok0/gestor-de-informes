from openpyxl import load_workbook, Workbook
from movimiento import Movimiento
import datetime
from openpyxl.styles import Font

class Lector:
    def __init__(self) -> None:
        self.wb = ''
        self.movimientos = []
    
    def algoritmo_tabla_1(self, row, fecha):
        if row[2] != None:
            if not str(row[2]).startswith('=') and row[2] != ' ':
                mov1 = {
                    'tipo': 'EGRESO',
                    'fecha': fecha,
                    'categoria': row[1],
                    'met_pago': '-',
                    'monto': row[2],
                    'sector': 'SAN RAFAEL',
                    'obs': ''
                }
                self.movimientos.append(Movimiento(mov=mov1))

        if row[3] != None:
            if not str(row[3]).startswith('=') and row[3] != ' ':
                mov2 = {
                    'tipo': 'EGRESO',
                    'fecha': fecha,
                    'categoria': row[1],
                    'met_pago': '-',
                    'monto': row[3],
                    'sector': 'ALVEAR',
                    'obs': ''
                }
                self.movimientos.append(Movimiento(mov=mov2))

        if row[4] != None:
            if not str(row[4]).startswith('=')  and row[4] != ' ':
                mov3 = {
                    'tipo': 'EGRESO',
                    'fecha': fecha,
                    'categoria': row[1],
                    'met_pago': '-',
                    'monto': row[4],
                    'sector': 'SAN MARTÍN',
                    'obs': ''
                }
                self.movimientos.append(Movimiento(mov=mov3))
            
        return self.movimientos

    def algoritmo_tabla_2(self, row, fecha):
        if row[2] != None:
            if not str(row[2]).startswith('=') and row[2] != ' ':
                mov1 = {
                    'tipo': 'EGRESO',
                    'fecha': fecha,
                    'categoria': row[1],
                    'met_pago': '-',
                    'monto': row[2],
                    'sector': 'GRANJA',
                    'obs': ''
                }
                self.movimientos.append(Movimiento(mov=mov1))
        if row[4] != None:
            if not str(row[4]).startswith('=') and row[2] != ' ':
                mov2 = {
                    'tipo': 'EGRESO',
                    'fecha': fecha,
                    'categoria': row[3],
                    'met_pago': '-',
                    'monto': row[4],
                    'sector': 'CAMPO',
                    'obs': ''
                }
                self.movimientos.append(Movimiento(mov=mov2))
        return self.movimientos

    
    def algoritmo_tabla_3(self, row, fecha):
        if row[2] != None:
            if not str(row[2]).startswith('=') and row[2] != ' ':
                mov1 = {
                    'tipo': 'EGRESO',
                    'fecha': fecha,
                    'categoria': row[1],
                    'met_pago': '-',
                    'monto': row[2],
                    'sector': 'FRIGO',
                    'obs': ''
                }
                self.movimientos.append(Movimiento(mov=mov1))
        if row[4] != None:
            if not str(row[4]).startswith('=') and row[4] != ' ':
                mov2 = {
                    'tipo': 'EGRESO',
                    'fecha': fecha,
                    'categoria': row[3],
                    'met_pago': '-',
                    'monto': row[4],
                    'sector': 'ADMIN',
                    'obs': ''
                }
                self.movimientos.append(Movimiento(mov=mov2))
        return self.movimientos

    def cargar_consolidacion_gastos(self, ruta, fecha):
        self.wb = load_workbook(ruta)
        sheet = self.wb.active
        # Hay 3 algoritmos posibles para guardar los datos según la tabla en la que se esté. 
        # Se ejecutará uno u otro algoritmo según el encabezado que se haya encontrado.
        algoritmo = 'TABLA_1'
        for row in sheet.iter_rows(values_only=True):
            if row[2] == 'GRANJA':
                algoritmo = 'TABLA_2'
            elif row[2] == 'FRIGORÍFICO':
                algoritmo = 'TABLA_3'
            
            if row[2] != 'LOCAL SAN RAFAEL' and row[2] != 'GRANJA' and row[2] != 'FRIGORÍFICO':    
                if algoritmo == 'TABLA_1':
                    self.algoritmo_tabla_1(row, fecha)
                elif algoritmo == 'TABLA_2':
                    self.algoritmo_tabla_2(row, fecha)
                elif algoritmo == 'TABLA_3':
                    self.algoritmo_tabla_3(row, fecha)
        
        return self.movimientos

    def cargar_caja_chica(self, ruta, sucursal, met_pago):
        # Se carga el WORKBOOK de excel que se va a usar y se activa la primer hoja
        self.wb = load_workbook(ruta)
        sheet = self.wb.active
        # La lista "movimientos" será aquella en la que guardemos todos los movimientos. 
        # Esta será devuelta al final de la función

        # Hacemos corresponder a cada mes su número para luego formatear bien la fecha
        meses = {
            'Jan':'01',
            'Feb':'02',
            'Mar':'03',
            'Apr':'04',
            'May':'05',
            'Jun':'06',
            'Jul':'07',
            'Aug':'08',
            'Sep':'09',
            'Oct':'10',
            'Nov':'11',
            'Dec':'12',
        }

        # Por cada fila leída del excel...
        for row in sheet.iter_rows(values_only=True):
            
            # Este IF está para asegurarse de que no estamos tomando una fila llena de valores nulos, ni 
            # estamos tomando los ingresos. Tampoco estamos tomando la primer columna que tiene los 
            # encabezados de las columnas, cuyo primer valor es "fecha".
            if row[0] != None and row[4] != 0.0 and row[0] != 'fecha' and not 'EGRESO' in row[1] and not 'INGRESO' in row[1] and not 'SUELDO' in row[1] and not 'RETIRO' in row[1]:
                mov = {
                'tipo': 'EGRESO',
                'fecha': '',
                'categoria': '',
                'met_pago': met_pago,
                'monto': '',
                'sector': '',
                'obs': ''
                }
                if 'FRIGO' in row[2]:
                    mov['sector'] = 'FRIGO'
                elif 'GRANJA' in row[2]:
                    mov['sector'] = 'GRANJA'
                elif 'ADMIN' in row[2]:
                    mov['sector'] = 'ADMIN'
                else:
                    mov['sector'] = sucursal
                mov['fecha'] = row[0][:2] +'-'+ meses[row[0][3:6]] + '-20' + row[0][7:]
                mov['categoria'] = row[1]
                mov['obs'] = row[2]
                mov['monto'] = row[4]
                # Creamos el movimiento con los datos extraídos
                self.movimientos.append(Movimiento(mov=mov))

        return self.movimientos
    
    def escribir_informe(self, movimientos, fecha_inicio, fecha_fin):
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe" + datetime.datetime.now().strftime('%Y-%m-%d')

        # Escribir encabezados
        encabezados = ["Tipo", "Fecha", "Categoría", "Método de Pago", "Monto", "Sector", "Observaciones"]
        ws.append(encabezados)
        
        # Dar formato a los encabezados
        for cell in ws[1]:
            cell.font = Font(bold=True)

        
        # Filtrado de los datos
        movimientos_filtrado = []
        mov_fecha = {}
        if fecha_inicio == None and fecha_fin == None:
            movimientos_filtrado = movimientos
        else:
            fecha_inicio = datetime.datetime.strptime(fecha_inicio, '%d-%m-%Y')
            fecha_fin = datetime.datetime.strptime(fecha_fin, '%d-%m-%Y')
            for movimiento in movimientos:    
                fecha = datetime.datetime.strptime(movimiento.fecha, '%d-%m-%Y')
                mov_fecha[movimiento] = fecha
                if fecha_inicio <= fecha <= fecha_fin:
                    movimientos_filtrado.append(movimiento)

        # Copiar datos de los objetos Movimiento
        for movimiento in movimientos_filtrado:
            fila = [
                movimiento.tipo,
                movimiento.fecha,
                movimiento.categoria,
                movimiento.met_pago,
                movimiento.monto,
                movimiento.sector,
                movimiento.obs,
                
            ]
            ws.append(fila)
        
        # Guardar el archivo
        wb.save(ws.title + '.xlsx')

    def cargar_informe(self, ruta_excel):
        # Cargar el archivo de Excel
        wb = load_workbook(ruta_excel)
        sheet = wb.active  # Utilizar la hoja activa

        movimientos = []

        # Iterar para obtener los valores del excel

        for row in sheet.iter_rows(values_only=True):
            if row[0] != 'Tipo':
                mov = {
                    'tipo': row[0],
                    'fecha': row[1],
                    'categoria': row[2],
                    'met_pago': row[3],
                    'monto': row[4],
                    'sector': row[5],
                    'obs': row[6]
                }

                movimientos.append(Movimiento(mov))
        
        return movimientos

    def cargar_registro_contrable(self, ruta, sheetname):
        self.wb = load_workbook(ruta)
        sheet = self.wb[sheetname]
        
        movimientos = []

        for row in sheet.iter_rows(values_only=True, min_row=9):
            if row[3] != None:
                
                print(row)
                if sheetname == 'Ingresos':
                    
                    mov = {
                        'fecha': row[1].strftime('%Y-%m-%d'),
                        'categoria': row[3],
                        'met_pago': row[5],
                        'monto': row[6],
                        'obs': ''
                    }
                    
                    mov['tipo'] = 'INGRESO'

                    if 'Alvear' in row[3]:
                        mov['sector'] = 'ALVEAR'
                    elif 'San Martín' in row[3]:
                        mov['sector'] = 'SAN MARTIN'
                    elif 'San Rafael' in row[3]:
                        mov['sector'] = 'SAN RAFAEL'
                    elif 'Casa Central' in row[3] or 'Casa central' in row[3]:
                        mov['sector'] = 'ADMIN'
                    elif 'Granja' in row[3] or 'granja' in row[3] or 'Campo' in row[3] or 'campo' in row[3]:
                        mov['sector'] = 'GRANJA'
                    elif 'Frigorífico' in row[3] or 'Frigorifico' in row[3]:
                        mov['sector'] = 'FRIGO'
                    else:
                        mov['sector'] = 'Asignar'
                else:
                    
                    mov = {
                        'fecha': row[1],
                        'categoria': row[3],
                        'met_pago': row[5], #7?
                        'monto': row[10],
                        'obs': row[4]
                    }
                    if type(row[1]) == datetime.datetime:
                        mov['fecha'] = row[1].strftime('%Y-%m-%d')
                    mov['tipo'] = 'EGRESO'

                    if 'Alvear' in row[4]:
                        mov['sector'] = 'ALVEAR'
                    elif 'San Martín' in row[4]:
                        mov['sector'] = 'SAN MARTIN'
                    elif 'San Rafael' in row[4]:
                        mov['sector'] = 'SAN RAFAEL'
                    elif 'Casa Central' in row[4] or 'Casa central' in row[4]:
                        mov['sector'] = 'ADMIN'
                    elif 'Granja' in row[4] or 'granja' in row[4] or 'Campo' in row[4] or 'campo' in row[4]:
                        mov['sector'] = 'GRANJA'
                    elif 'Frigorífico' in row[4] or 'Frigorifico' in row[4]:
                        mov['sector'] = 'FRIGO'
                    else:
                        mov['sector'] = 'Asignar'
                movimientos.append(Movimiento(mov))
        return movimientos

